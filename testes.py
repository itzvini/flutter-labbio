from enum import _auto_null
from cv2 import COLOR_BGR2GRAY
import numpy as np
import cv2 as cv
import matplotlib.pyplot as plt
import openpyxl
import time
#Abre o Arquivo de Excel onde serão salvos os dados
wb = openpyxl.load_workbook('teste.xlsx')
ws = wb.active #Abre a aba ativa do arquivo de excel

#Função para associar um numero de iteração a uma coluna do excel
def find_letter(iterador):
    if iterador == 0:
        return 'A'
    if iterador == 1:
        return 'B'
    if iterador == 2:
        return 'C'
    if iterador == 3:
        return 'D'
    if iterador == 4:
        return 'E'
    if iterador == 5:
        return 'F'
    if iterador == 6:
        return 'G'
    if iterador == 7:
        return 'H'
    if iterador == 8:
        return 'I'
    if iterador == 9:
        return 'J'
    if iterador == 10:
        return 'K'
    if iterador == 11:
        return 'L'
    if iterador == 12:
        return 'M'
    if iterador == 13:
        return 'N'
    if iterador == 14:
        return 'O'
    if iterador == 15:
        return 'P'
    if iterador == 16:
        return 'Q'
    if iterador == 17:
        return 'R'
    if iterador == 18:
        return 'S'
    if iterador == 19:
        return 'T'
    if iterador == 20:
        return 'U'
    if iterador == 21:
        return 'V'
    if iterador == 22:
        return 'W'
    if iterador == 23:
        return 'X'
    if iterador == 24:
        return 'Y'
    if iterador == 25:
        return 'Z'
    if iterador == 26:
        return 'AA'
    if iterador == 27:
        return 'AB'
    if iterador == 28:
        return 'AC'
    if iterador == 29:
        return 'AD'
    if iterador == 30:
        return 'AE'
    if iterador == 31:
        return 'AF'
    if iterador == 32:
        return 'AG'
    if iterador == 33:
        return 'AH'
    if iterador == 34:
        return 'AI'
    if iterador == 35:
        return 'AJ'
    if iterador == 36:
        return 'AK'
    if iterador == 37:
        return 'AL'
    if iterador == 38:
        return 'AM'
    if iterador == 39:
        return 'AN'
    if iterador == 40:
        return 'AO'
    if iterador == 41:
        return 'AP'
    if iterador == 42:
        return 'AQ'
    if iterador == 43:
        return 'AR'
    if iterador == 44:
        return 'AS'
    if iterador == 45:
        return 'AT'
    if iterador == 46:
        return 'AU'
    if iterador == 47:
        return 'AV'
    if iterador == 48:
        return 'AW'
    if iterador == 49:
        return 'AX'
    if iterador == 50:
        return 'AY'

#Função para apagar todas as linhas do excel
def delete_all_rows():
    while ws['A1'].value != None:
        ws.delete_rows(1)
    return
    
#Função para achar a borda mais proxima do centro da imagem
def find_nearest_white(canny, target):
    nonzero = cv.findNonZero(canny)
    distances = np.sqrt((nonzero[:,:,0] - target[0]) ** 2 + (nonzero[:,:,1] - target[1]) ** 2)
    nearest_index = np.argmin(distances)
    return nonzero[nearest_index]

#Obtem o nome do arquivo de imagem
img = cv.VideoCapture('7.mp4')
prevCircle = None
dist = lambda x1, y1, x2, y2: np.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2)
#Define o tamanho da imagem
frame_width = int(img.get(3))
frame_height = int(img.get(4))
size = (frame_width, frame_height)

#Iniciador onde será salvo o video
result = cv.VideoWriter('00.mp4', cv.VideoWriter_fourcc(*'mp4v'), 10, size)
                 
#Input para receber a quantidade de linhas que aparecerão em cada folheto
number_of_angles = int(input("Digite o numero de linhas para cada folheto: "))

#Inicializ os dados do Excel vazios
delete_all_rows()

#Variavel para auxiliar na manipulação de linhas e colunas do excel
count = 2

#Variavel de linhas totais sendo igual ao numero de angulos multiplicado pelo nuemro de folhetos
linhas = number_of_angles*3

#Variavel de Angulo
angulo = 360/linhas

#Parte que define quantas colunas serão necessarias para o excel
iterador = 1
for iterador in range(linhas+1):
    letra = find_letter(iterador)
    ws[f'{letra}1'] = angulo*iterador

#Loop principal de iteracao
while True:
    ws[f'A{count}'] = str(count-1) 
    ret, frame = img.read()
    if not ret:
        break
    
    w0 = frame.shape[0]
    h0 = frame.shape[1]
    
    # Ellipse parameters
    radius = 145
    axes = (radius, radius)
    angle = 0
    
    thickness = -1
    back = cv.imread('white.png')
    
    W = (255, 255, 255) #Branco
    Y = (127, 255, 212) #Amarelo
    R = 0, 0, 255 #Vermelho
    G = 0, 255, 0 #Verde
    B = 255, 0, 0 #Azul
    
    
    # grayFrame = cv.cvtColor(frame, cv.COLOR_BGR2GRAY)
    # blurFrame = cv.GaussianBlur(grayFrame, (5, 5), 0)
    # circles = cv.HoughCircles(blurFrame, cv.HOUGH_GRADIENT, 1.4, 1000, param1=100, param2=200, minRadius=143, maxRadius=147)
    # if circles is not None:
    #     circles = np.uint16(np.around(circles))
    #     chosen = None
    #     for i in circles[0, ]:
    #         if chosen is None: chosen = i
    #         if prevCircle is not None:
    #             if dist(chosen[0],chosen[1],prevCircle[0],prevCircle[1]) <= dist(i[0],i[1],prevCircle[0],prevCircle[1]):
    #                 chosen = i
    #     cv.circle(frame, (chosen[0], chosen[1]), chosen[2], (0, 255, 0), 2)
    #     prevCircle = chosen
    
    # centro = (chosen[0], chosen[1])
    centro = [307, 250]
    loop = 0
    i = 1
    k = 0
    iterador = 1
    imgcopy = frame.copy()
    while loop < linhas:
        mask = np.ones((w0,h0),dtype=np.uint8)
        src1_mask=cv.cvtColor(mask,cv.COLOR_GRAY2BGR)#change mask to a 3 channel image
        letra = find_letter(iterador)
        if cv.waitKey(1) & 0xFF == ord('q'):
            img.release()
            result.release()
            cv.destroyAllWindows()
        startAngle = (k*angulo)
        endAngle = (i*angulo)
        if loop == linhas-1:
            endAngle = 360
        mascara = cv.inRange(src1_mask, (0, 0, 50), (20, 20,255))
        cv.ellipse(src1_mask, centro, axes, angle, startAngle, endAngle, (255,255,255), thickness)
        mask_out=cv.subtract(src1_mask, imgcopy)
        mask_out=cv.subtract(src1_mask,mask_out)
        
        canny = cv.inRange(mask_out, (0, 0, 50), (50, 50,255))
        try:
            pointAB = find_nearest_white(canny, (centro))
            if startAngle >= 0 and endAngle <= 120:
                color = B
            if startAngle >= 120 and endAngle <= 240:
                color = G
            if startAngle >= 240 and endAngle <= 360:
                color = Y
            cv.line(frame, (centro), (int(pointAB[0][0]), int(pointAB[0][1])), (color), 1)
            raio = np.sqrt((pointAB[0][0] - centro[0]) ** 2 + (pointAB[0][1] - centro[1]) ** 2)
            print(startAngle, endAngle, f'raio = {raio}')
            ws[f'{letra}{count}'] = raio
        except:
            print(f"nenhum ponto branco em {startAngle} e {endAngle}")
        iterador += 1
        i += 1
        k += 1
        loop += 1
        
    
    count += 1
    cv.imshow('mask', frame)   
    result.write(frame)
    wb.save('teste.xlsx')
    #takes a screenshot
    if cv.waitKey(1) == ord('p'):
        cv.imwrite('screenshot.png', frame)

    if cv.waitKey(1) & 0xFF == ord('q'):
        break
    
img.release()
result.release()
cv.destroyAllWindows()